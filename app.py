from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from datetime import datetime, timezone, timedelta
from functools import wraps
import glob
import openpyxl
import os
import time

from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")

APP_PASSWORD = os.environ.get("APP_PASSWORD", "")


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if APP_PASSWORD and not session.get("authenticated"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

ONEDRIVE_SHARE_URL = os.environ.get("ONEDRIVE_SHARE_URL", "").strip()
def _find_xlsm():
    pattern = os.path.join(os.path.dirname(__file__), "..", "ebayマスタ*.xlsm")
    matches = glob.glob(pattern)
    return matches[0] if matches else None

XLSM_PATH = _find_xlsm()
SHEET_NAME = "仕入・在庫管理表"
HEADER_ROW = 5
DATA_START_ROW = 6
MAX_COL = 28  # A〜AB

# キャッシュ: {rows_data: {行番号: [値リスト]}, headers: [...], mtime, source, max_row}
_cache = {"rows_data": None, "headers": None, "mtime": None, "source": None, "max_row": 0}


def _parse_workbook(wb):
    """workbook から全行データとヘッダーを抽出してメモリにキャッシュする形式で返す"""
    ws = wb[SHEET_NAME]

    # read_only モードでは iter_rows で最終行まで一括読み込み（max_row 指定なし）
    rows_data = {}
    max_data_row = HEADER_ROW
    for row in ws.iter_rows(min_row=HEADER_ROW, max_col=MAX_COL, values_only=False):
        r = row[0].row
        values = [cell.value for cell in row]
        # 全セルが空の行はスキップ
        if any(v is not None for v in values):
            rows_data[r] = values
            max_data_row = r

    # ヘッダー構築
    header_values = rows_data.get(HEADER_ROW, [None] * MAX_COL)
    headers = []
    for col in range(MAX_COL):
        col_letter = openpyxl.utils.get_column_letter(col + 1)
        val = header_values[col] if col < len(header_values) else None
        headers.append({"col": col + 1, "letter": col_letter, "name": val or f"({col_letter})"})

    wb.close()
    return rows_data, headers, max_data_row


def get_data():
    """行データとヘッダーを返す（キャッシュ付き）"""
    if ONEDRIVE_SHARE_URL:
        return _get_data_onedrive()
    return _get_data_local()


def _get_data_onedrive():
    """OneDrive 版: read_only + メモリキャッシュ"""
    from onedrive import fetch_xlsm, _is_cache_fresh

    if _cache["source"] == "onedrive" and _cache["rows_data"] is not None and _is_cache_fresh():
        return _cache["rows_data"], _cache["headers"]

    buf = fetch_xlsm(ONEDRIVE_SHARE_URL)
    wb = openpyxl.load_workbook(buf, data_only=True, keep_vba=False, read_only=True)
    rows_data, headers, max_row = _parse_workbook(wb)
    _cache.update(rows_data=rows_data, headers=headers, mtime=None, source="onedrive", max_row=max_row)
    return rows_data, headers


def _refresh_onedrive():
    """OneDrive キャッシュを強制クリアして再取得"""
    from onedrive import invalidate_cache
    invalidate_cache()
    _cache.update(rows_data=None, headers=None, mtime=None, source=None)


def _get_data_local():
    """ローカルファイル版: mtime ベースのキャッシュ"""
    global XLSM_PATH
    if not XLSM_PATH:
        XLSM_PATH = _find_xlsm()
    if not XLSM_PATH:
        raise FileNotFoundError("ebayマスタ*.xlsm が見つかりません")
    mtime = os.path.getmtime(XLSM_PATH)
    if _cache["mtime"] != mtime or _cache["source"] != "local":
        wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, keep_vba=False, read_only=True)
        rows_data, headers, max_row = _parse_workbook(wb)
        _cache.update(rows_data=rows_data, headers=headers, mtime=mtime, source="local", max_row=max_row)
    return _cache["rows_data"], _cache["headers"]


@app.route("/login", methods=["GET", "POST"])
def login():
    if not APP_PASSWORD:
        return redirect(url_for("index"))
    if request.method == "POST":
        if request.form.get("password") == APP_PASSWORD:
            session["authenticated"] = True
            return redirect(url_for("index"))
        return render_template("login.html", error="パスワードが違います")
    return render_template("login.html")


@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/api/refresh", methods=["POST"])
@login_required
def api_refresh():
    """キャッシュをクリアして最新データを再取得する"""
    if ONEDRIVE_SHARE_URL:
        _refresh_onedrive()
        t0 = time.time()
        get_data()
        elapsed = time.time() - t0
        return jsonify({"status": "refreshed", "source": "onedrive", "elapsed_s": round(elapsed, 2)})
    return jsonify({"status": "skipped", "source": "local", "message": "ローカルモードではmtimeで自動更新されます"})


@app.route("/api/fileinfo")
@login_required
def api_fileinfo():
    """取得したファイルの情報を返す（デバッグ用）"""
    JST = timezone(timedelta(hours=9))
    if ONEDRIVE_SHARE_URL:
        from onedrive import get_file_info
        info = get_file_info()
        fetched_at = None
        if info["fetched_at"]:
            fetched_at = datetime.fromtimestamp(info["fetched_at"], tz=JST).strftime("%Y-%m-%d %H:%M:%S JST")
        return jsonify({
            "source": "onedrive",
            "filename": info["filename"],
            "fetched_at": fetched_at,
            "content_length": info["content_length"],
            "content_type": info.get("content_type"),
            "final_url": info.get("final_url", "")[:100] + "...",
            "status_code": info.get("status_code"),
            "share_url": ONEDRIVE_SHARE_URL[:50] + "...",
        })
    else:
        fname = os.path.basename(XLSM_PATH) if XLSM_PATH else None
        mtime = None
        if XLSM_PATH and os.path.exists(XLSM_PATH):
            mtime = datetime.fromtimestamp(os.path.getmtime(XLSM_PATH), tz=JST).strftime("%Y-%m-%d %H:%M:%S JST")
        return jsonify({
            "source": "local",
            "filename": fname,
            "last_modified": mtime,
            "path": XLSM_PATH,
        })


@app.route("/api/headers")
@login_required
def api_headers():
    _, headers = get_data()
    return jsonify(headers)


@app.route("/api/cell")
@login_required
def get_cell():
    try:
        col = int(request.args.get("col", 1))
        row = int(request.args.get("row", DATA_START_ROW))
    except (TypeError, ValueError):
        return jsonify({"error": "col と row は整数で指定してください"}), 400

    if col < 1 or col > MAX_COL:
        return jsonify({"error": f"col は 1〜{MAX_COL} の範囲で指定してください"}), 400
    rows_data, headers = get_data()
    max_row = _cache["max_row"]
    if row < DATA_START_ROW or row > max_row:
        return jsonify({"error": f"row は {DATA_START_ROW}〜{max_row} の範囲で指定してください"}), 400
    row_values = rows_data.get(row)
    value = row_values[col - 1] if row_values and col - 1 < len(row_values) else None
    return jsonify({
        "col": col,
        "row": row,
        "header": headers[col - 1]["name"],
        "value": str(value) if value is not None else None,
    })


@app.route("/api/search")
@login_required
def api_search():
    """出品管理ID（B列）で行を検索し、該当行の全データを返す"""
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify({"error": "検索IDを入力してください"}), 400

    rows_data, headers = get_data()

    matched = []
    for r in range(DATA_START_ROW, _cache["max_row"] + 1):
        row_values = rows_data.get(r)
        if not row_values:
            continue
        b_val = row_values[1]  # B列 = index 1
        if b_val is None:
            continue
        # 数値の場合 12345.0 → "12345" に変換
        b_str = str(int(b_val)) if isinstance(b_val, float) and b_val == int(b_val) else str(b_val)
        if query in b_str.strip():
            row_data = {}
            for col in range(MAX_COL):
                val = row_values[col] if col < len(row_values) else None
                row_data[headers[col]["letter"]] = str(val) if val is not None else None
            matched.append({"row": r, "data": row_data})

    return jsonify({
        "query": query,
        "count": len(matched),
        "headers": headers,
        "rows": matched,
    })


@app.route("/api/range")
@login_required
def get_range():
    try:
        row_start = int(request.args.get("row_start", DATA_START_ROW))
        row_end = int(request.args.get("row_end", DATA_START_ROW + 19))
    except (TypeError, ValueError):
        return jsonify({"error": "row_start と row_end は整数で指定してください"}), 400

    row_start = max(row_start, DATA_START_ROW)
    rows_data, headers = get_data()
    row_end = min(row_end, _cache["max_row"])

    if row_end - row_start > 100:
        row_end = row_start + 100

    rows = []
    for r in range(row_start, row_end + 1):
        row_data = {}
        row_values = rows_data.get(r)
        for col in range(MAX_COL):
            val = row_values[col] if row_values and col < len(row_values) else None
            row_data[headers[col]["letter"]] = str(val) if val is not None else None
        rows.append({"row": r, "data": row_data})

    return jsonify({
        "row_start": row_start,
        "row_end": row_end,
        "headers": headers,
        "rows": rows,
    })


if __name__ == "__main__":
    source = "OneDrive" if ONEDRIVE_SHARE_URL else f"Local ({XLSM_PATH})"
    print(f"Source: {source}")
    print(f"Sheet: {SHEET_NAME}")
    app.run(debug=True, host="0.0.0.0", port=5000)
